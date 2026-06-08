import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

data = [
    ["分类", "字段名称 (Field)", "说明 / 示例", "产品01 (示例)", "产品02", "产品03", "产品04", "产品05"],
    ["A. 基本信息", "Product Name (商品名称)", "建议128字符内，包含核心词+修饰词", "Cotton Breathable T-Shirt", "", "", "", ""],
    ["A. 基本信息", "Keyword 1 (关键词1)", "主推核心引流词", "T-Shirt", "", "", "", ""],
    ["A. 基本信息", "Keyword 2 (关键词2)", "精准搜索词", "Cotton T-Shirt", "", "", "", ""],
    ["A. 基本信息", "Keyword 3 (关键词3)", "长尾搜索词", "Men Breathable Summer T-Shirt", "", "", "", ""],
    ["A. 基本信息", "Product Group (商品分组)", "对应阿里后台的分组名称", "Apparel / T-Shirts", "", "", "", ""],
    ["B. 属性信息", "Origin (产地)", "如：Hebei, China", "Zhejiang, China", "", "", "", ""],
    ["B. 属性信息", "Brand (品牌)", "如：OEM / 自有品牌", "OEM", "", "", "", ""],
    ["B. 属性信息", "Model Number (型号)", "如：产品具体型号", "TS-2026-01", "", "", "", ""],
    ["B. 属性信息", "Material (材质)", "如：Q235 Steel / 100% Cotton", "100% Cotton", "", "", "", ""],
    ["B. 属性信息", "Application (应用)", "主要应用场景", "Daily Wear / Summer / Sports", "", "", "", ""],
    ["B. 属性信息", "Feature (特性)", "如：Eco-Friendly, Durable", "Breathable, Anti-Pilling, Soft", "", "", "", ""],
    ["C. 自定义属性", "Attribute 1 - Name", "自定义属性名（选填）", "Style", "", "", "", ""],
    ["C. 自定义属性", "Attribute 1 - Value", "自定义属性值（选填）", "Casual", "", "", "", ""],
    ["C. 自定义属性", "Attribute 2 - Name", "自定义属性名", "Sleeve Length", "", "", "", ""],
    ["C. 自定义属性", "Attribute 2 - Value", "自定义属性值", "Short Sleeve", "", "", "", ""],
    ["C. 自定义属性", "Attribute 3 - Name", "自定义属性名", "Collar", "", "", "", ""],
    ["C. 自定义属性", "Attribute 3 - Value", "自定义属性值", "O-Neck", "", "", "", ""],
    ["D. 交易信息", "MOQ (最小起订量)", "如：10 / 100", "50", "", "", "", ""],
    ["D. 交易信息", "Unit (计量单位)", "如：Pieces / Sets", "Pieces", "", "", "", ""],
    ["D. 交易信息", "FOB Price (FOB单价)", "如：5.5 (USD)", "4.5", "", "", "", ""],
    ["D. 交易信息", "Lead Time (交期)", "如：15 days", "15 days", "", "", "", ""],
    ["E. 物流信息", "Package Weight (包装重量)", "单位通常为kg，如：1.5", "0.2", "", "", "", ""],
    ["E. 物流信息", "Package Size (包装尺寸)", "长x宽x高 cm，如：30x20x10", "20x15x2", "", "", "", ""],
    ["F. 媒体资源", "Main Image URL (主图链接)", "第一张为首图", "https://example.com/img1.jpg", "", "", "", ""],
    ["F. 媒体资源", "Detail Image URLs (详情图)", "多张图可逗号分隔", "https://example.com/img2.jpg", "", "", "", ""],
    ["F. 媒体资源", "Video URL (视频链接)", "主图视频链接", "", "", "", "", ""],
    ["G. 详细描述", "Detail Description (详情描述)", "AI生成的HTML源码或文案", "<p>High quality 100% cotton T-shirt...</p>", "", "", "", ""],
]

df = pd.DataFrame(data[1:], columns=data[0])
file_path = r"C:\Users\57673\Desktop\阿里巴巴批量发品模板_竖版(参考MIC).xlsx"

with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name="批量发品模板", index=False)
    workbook = writer.book
    worksheet = writer.sheets["批量发品模板"]
    
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        
    example_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    for row in range(2, len(data) + 1):
        worksheet.cell(row=row, column=4).fill = example_fill
        
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    worksheet.column_dimensions['A'].width = 15
    worksheet.column_dimensions['B'].width = 30
    worksheet.column_dimensions['C'].width = 35
    for col in ['D', 'E', 'F', 'G', 'H']:
        worksheet.column_dimensions[col].width = 40
        
    for row in worksheet.iter_rows(min_row=1, max_row=len(data), min_col=1, max_col=len(data[0])):
        for cell in row:
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = border
            
    worksheet.insert_rows(1)
    worksheet.merge_cells('A1:H1')
    title_cell = worksheet['A1']
    title_cell.value = "Alibaba（国际站）批量发品信息表  |  Product Listing Batch Template (参考MIC排版格式)"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

    ws2 = workbook.create_sheet(title="使用说明")
    instructions = [
        ["Alibaba 竖版发品表 — 使用说明"],
        ["字段", "说明"],
        ["Step 1", "打开「批量发品模板」Sheet，每一列对应一个产品（D=产品01，E=产品02，F=产品03...）。"],
        ["Step 2", "黄色列区域为示例，白底色区域为待填区域，让AI生成后竖向粘贴至产品列。"],
        ["Step 3", "影刀RPA读取时：使用“循环Excel列”功能，每一列即为一个商品的全部属性，方便映射抓取。"],
        ["注意事项", "说明"],
        ["AI 生成提示词", "请向AI提供专用提示词（见聊天框），AI将输出用换行符分隔的一列数据，您可以直接从第一行复制粘贴到Excel对应的一列中。"]
    ]
    for row_idx, row in enumerate(instructions, 1):
        ws2.append(row)
        if row_idx == 1:
            ws2.merge_cells('A1:B1')
            ws2['A1'].font = Font(bold=True, size=14)
        elif row_idx in [2, 6]:
            for cell in ws2[row_idx]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
                
    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 90
    for row in ws2.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='center', wrap_text=True)

print("FILE_CREATED")